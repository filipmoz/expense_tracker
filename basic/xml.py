"""Generate a sample .xlsx from this directory. Run: python xml.py"""
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path

out = Path(__file__).resolve().parent / "output_openpyxl.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "Data"
ws["A1"] = "Name"
ws["B1"] = "Value"
ws["A1"].font = Font(bold=True)
ws["A2"], ws["B2"] = "Example", 42
wb.save(out)
print("Wrote", out)
