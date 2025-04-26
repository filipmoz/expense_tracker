"""
Excel export functionality using openpyxl and xlsxwriter
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xlsxwriter import Workbook as XlsxWriterWorkbook
from typing import List, Dict
from datetime import datetime
from io import BytesIO
from sqlalchemy.orm import Session
from app.models import Expense
from app.statistics import calculate_expense_statistics

def export_expenses_to_excel(db: Session) -> BytesIO:
    """Export expenses to Excel with data and charts"""
    expenses = db.query(Expense).order_by(Expense.date.desc()).all()
    stats = calculate_expense_statistics(db)
    
    # Create in-memory file
    output = BytesIO()
    
    # Note: xlsxwriter doesn't support reading from BytesIO directly in the way we need
    # So we'll create a temporary approach
    import tempfile
    import os
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = temp_file.name
    temp_file.close()
    
    # Use xlsxwriter for charts (write to temp file first)
    workbook = XlsxWriterWorkbook(temp_path)
    
    # Data sheet
    worksheet = workbook.add_worksheet('Expenses')
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#366092',
        'font_color': 'white',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd hh:mm:ss'})
    currency_format = workbook.add_format({'num_format': '£#,##0.00'})
    
    # Headers
    headers = ['ID', 'Date', 'Category', 'Description', 'Amount']
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header, header_format)
    
    # Data rows
    for row_num, expense in enumerate(expenses, 1):
        worksheet.write(row_num, 0, expense.id)
        worksheet.write_datetime(row_num, 1, expense.date, date_format)
        worksheet.write(row_num, 2, expense.category)
        worksheet.write(row_num, 3, expense.description or '')
        worksheet.write(row_num, 4, expense.amount, currency_format)
    
    # Auto-adjust column widths
    worksheet.set_column('A:A', 8)  # ID
    worksheet.set_column('B:B', 20)  # Date
    worksheet.set_column('C:C', 15)  # Category
    worksheet.set_column('D:D', 30)  # Description
    worksheet.set_column('E:E', 12)  # Amount
    
    # Statistics sheet
    stats_sheet = workbook.add_worksheet('Statistics')
    
    stats_headers = ['Metric', 'Value']
    for col_num, header in enumerate(stats_headers):
        stats_sheet.write(0, col_num, header, header_format)
    
    stats_data = [
        ['Total Expenses', stats['total_expenses']],
        ['Total Count', stats['total_count']],
        ['Average Expense', stats['average_expense']],
        ['Median Expense', stats['median_expense']],
        ['Min Expense', stats['min_expense']],
        ['Max Expense', stats['max_expense']],
        ['Standard Deviation', stats['std_deviation']]
    ]
    
    for row_num, (metric, value) in enumerate(stats_data, 1):
        stats_sheet.write(row_num, 0, metric)
        if isinstance(value, float):
            stats_sheet.write(row_num, 1, value, currency_format if 'Expense' in metric else None)
        else:
            stats_sheet.write(row_num, 1, value)
    
    stats_sheet.set_column('A:A', 25)
    stats_sheet.set_column('B:B', 15)
    
    # Category breakdown sheet
    category_sheet = workbook.add_worksheet('Category Breakdown')
    
    cat_headers = ['Category', 'Total Amount', 'Count', 'Average']
    for col_num, header in enumerate(cat_headers):
        category_sheet.write(0, col_num, header, header_format)
    
    row_num = 1
    for category, total in stats['category_breakdown'].items():
        count = stats['category_counts'][category]
        avg = total / count if count > 0 else 0
        category_sheet.write(row_num, 0, category)
        category_sheet.write(row_num, 1, total, currency_format)
        category_sheet.write(row_num, 2, count)
        category_sheet.write(row_num, 3, avg, currency_format)
        row_num += 1
    
    category_sheet.set_column('A:A', 20)
    category_sheet.set_column('B:B', 15)
    category_sheet.set_column('C:C', 10)
    category_sheet.set_column('D:D', 15)
    
    # Create charts
    chart_sheet = workbook.add_worksheet('Charts')
    
    # Chart 1: Expenses by Category (Pie Chart)
    chart1 = workbook.add_chart({'type': 'pie'})
    chart1.add_series({
        'name': 'Expenses by Category',
        'categories': ['Category Breakdown', 1, 0, row_num - 1, 0],
        'values': ['Category Breakdown', 1, 1, row_num - 1, 1],
    })
    chart1.set_title({'name': 'Expenses by Category'})
    chart1.set_size({'width': 480, 'height': 300})
    chart_sheet.insert_chart('B2', chart1)
    
    # Chart 2: Category Counts (Bar Chart)
    chart2 = workbook.add_chart({'type': 'column'})
    chart2.add_series({
        'name': 'Number of Expenses',
        'categories': ['Category Breakdown', 1, 0, row_num - 1, 0],
        'values': ['Category Breakdown', 1, 2, row_num - 1, 2],
    })
    chart2.set_title({'name': 'Number of Expenses by Category'})
    chart2.set_x_axis({'name': 'Category'})
    chart2.set_y_axis({'name': 'Count'})
    chart2.set_size({'width': 480, 'height': 300})
    chart_sheet.insert_chart('B20', chart2)
    
    # Chart 3: Category Averages (Bar Chart)
    chart3 = workbook.add_chart({'type': 'column'})
    chart3.add_series({
        'name': 'Average Expense',
        'categories': ['Category Breakdown', 1, 0, row_num - 1, 0],
        'values': ['Category Breakdown', 1, 3, row_num - 1, 3],
    })
    chart3.set_title({'name': 'Average Expense by Category'})
    chart3.set_x_axis({'name': 'Category'})
    chart3.set_y_axis({'name': 'Amount (£)'})
    chart3.set_size({'width': 480, 'height': 300})
    chart_sheet.insert_chart('B38', chart3)
    
    workbook.close()
    
    # Read the file back into BytesIO
    with open(temp_path, 'rb') as f:
        output.write(f.read())
    
    # Clean up temp file
    os.unlink(temp_path)
    
    output.seek(0)
    return output
